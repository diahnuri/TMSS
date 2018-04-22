from django.test import TestCase
from django.shortcuts import render
import os
import numpy as np
import re,string
import json,csv
import sqlite3
from django.http import HttpResponse
from .models import *
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import TemporaryUploadedFile
from _io import StringIO, BytesIO
from wsgiref.util import FileWrapper
from django.http.response import HttpResponseRedirect
from zipfile import ZipFile
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.reader.excel import load_workbook
from jst.models import *
from dlnn.listFunction import *
from dlnn.listFunction2 import *
from dlnn.tests import *

# Create your tests here.

wmss_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
jst_file = os.path.join(wmss_folder, 'jst.models')

BASE_DIR = os.path.dirname((__file__))
var = os.path.join(BASE_DIR, 'var.txt')

def readfile(a):
    f = open(a,'r')
    c = f.read(1)
    f.close
    return c

class usedvar:
    abc = readfile(var)
    method = abc
    
def handle_uploaded_file(f, g):
    ##Jika ingin menyimpan dalam folder yang berbeda
    #filename, fileext = os.path.splitext(g)
    #g = "Media/"+fileext[1:]+"/"+g
#statis langsung
    g = "static/tables/"+g
    g = os.path.join(BASE_DIR, g)
    with open(g, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

def handle(f, g):
    nama, ext = os.path.splitext(g) 
    j = TemporaryUploadedFile(g,f.content_type,f.size,f.charset)
    for chunk in f.chunks():
        j.write(chunk)
    
    if ext.lower()==".json":
        savejson(j)
    elif ext.lower()==".csv":
        savecsv(j)
    elif ext.lower()==".xls" or ".xlsx":
        savexls(j)

def savejson(f):
    tss = json.loads(open(f.temporary_file_path()).read())
    for a in tss:
        name = a['NAME']
        date = a['DATE']
        tweet = a['TWEET']
        hashtags = a['HASHTAGS']
        RT = a['RETWEET_USER']
        q = TwitterCrawl(name=name,tweet = tweet, date= date, hashtag=hashtags,Retweet_user=RT)
        q.save()
def savecsv(f):
    f = open(f.temporary_file_path(), encoding="utf-8")
    reader = csv.reader(f)
    for row in reader:
        
#         row = unidecode(row)
#         print(row)
        name = row[0]
        tweet = row[1]
        date =row[2]
        RT =row[3]
        hashtags = row[4]
        q = TwitterCrawl(name=name,tweet = tweet, date= date, hashtag=hashtags,Retweet_user=RT)
        q.save()
    
    print()
    
def savexls(f):
#     j = f.read()
    wb = load_workbook(f)
    ws = wb.active
    for row in ws.iter_rows():
        name =str(row[0].value)
        tweet =str(row[1].value)
        date =str(row[2].value)[:10]
        RT =str(row[3].value)
        hashtags = str(row[4].value)
        q = TwitterCrawl(name=name,tweet = tweet, date= date, hashtag=hashtags,Retweet_user=RT)
        q.save()
        print('saved')
    
    print()
def getFile():
    ##jika ingin menggunakan media
    #json = os.path.join(BASE_DIR, "Media/JSON")
    #CommaSV = os.path.join(BASE_DIR, "Media/CSV")
    #a = os.listdir(json)
    #b = os.listdir(CommaSV)
#static langsung
    a = []
    b = []
    p = os.path.join(BASE_DIR,'static/tables/')
    data = os.listdir(p)
    for da in data:
        nama, ext = os.path.splitext(da) 
        if ext.lower()==".json":
            a.append(da)
        elif ext.lower()==".csv":
            b.append(da)
    return a,b

def getFilename():
    json, commasv = getFile()
    listnamaJSON = []
    listnamaCSV = []
    for a in json:
        a,b = os.path.splitext(a)
        listnamaJSON.append(a)
    for a in commasv:
        a,b = os.path.splitext(a)
        listnamaCSV.append(a)
    return listnamaJSON, listnamaCSV

def getKeyJson(data):
    p = os.path.join(BASE_DIR,'static/tables/')
    f = open(p+data, 'r', encoding='utf-8')
    g = json.load(f)
    c ={}
    b =[]
    for a in g:
        for key in a.keys():
            c[key]=''
    for key in c.keys():
        b.append(key)
    return b

def uploadtoarray(data):
    a = []
    for chunk in data.chunks():
        a.append(chunk)  
    return a

def getall():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_temp = os.path.join(BASE_DIR, 'db.sqlite3')
    
    conn = sqlite3.connect(db_temp)
    cursor = conn.cursor()
    a = cursor.execute("SELECT NAME, TWEET, RETWEET_USER, HASHTAG from social_media_crawling_TwitterCrawl")
    c = {'users':[],'twits':[],'ritwistUsr':[],'ritwits':[],'hashtags':[]}
    for row in a:
        c['users'].append(row[0])
        c['twits'].append(row[1])
        c['ritwits'].append(row[2])
        c['hashtags'].append(row[3])
    return c

def gettweet():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_temp = os.path.join(BASE_DIR, 'db.sqlite3')
    print(BASE_DIR)
    
    conn = sqlite3.connect(db_temp)
    cursor = conn.cursor()
    a = cursor.execute("SELECT TWEET from social_media_crawling_TwitterCrawl")
    c = []
    for row in a:
        c.append(row[0])
    conn.close()
    return c
        
def downloadfile(allfile, twitterdata):
    panjang = len(allfile)
    
    
    if panjang == 3:
        ##do something
        f = StringIO()    
        wb = Workbook()
        ws = wb.active
            
        for row in twitterdata:
            f.write(row.name+','+row.tweet+','+str(row.date)+','+str(row.Retweet_user)+','+row.hashtag)
            f.write('\n')
            ws.append([row.name,row.tweet,row.date,row.Retweet_user,row.hashtag])
            
        f.flush()
        f.seek(0)
        
        ##make json
        g = StringIO()
        data =[]
        for row in twitterdata:
            data1 = {}
            data1["NAME"]=row.name
            data1["TWEET"] = row.tweet
            data1["RETWEET_USER"] = row.Retweet_user
            data1["HASHTAGS"] = row.hashtag
            data1["DATE"]= str(row.date)
            data.append(data1)
        json.dump(data,g,indent=3)
        g.flush()
        g.seek(0)
        
        
        
        output = BytesIO()
        zip = ZipFile(output,'w')
        zip.writestr("Data_Twitter.csv",f.getvalue())
        zip.writestr("Data_Twitter.JSON",g.getvalue())
        zip.writestr("Data_Twitter.xlsx",save_virtual_workbook(wb))
        zip.close()
        response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
        response['Content-Disposition'] = 'attachment; filename=Data_Twitter.zip'
        
        return response
    
    elif panjang == 2:
        
        ##read list
        if '0' and '1' in allfile:
            ##return csv and json
            print('csv and jsn')
            #make csv
            f = StringIO()    
            
            for row in twitterdata:
                f.write(row.name+','+row.tweet+','+str(row.date)+','+str(row.Retweet_user)+','+row.hashtag)
                f.write('\n')
                
            f.flush()
            f.seek(0)
            
            ##make json
            g = StringIO()
            data =[]
            for row in twitterdata:
                data1 = {}
                data1["NAME"]=row.name
                data1["TWEET"] = row.tweet
                data1["RETWEET_USER"] = row.Retweet_user
                data1["HASHTAGS"] = row.hashtag
                data1["DATE"]= str(row.date)
                data.append(data1)
            json.dump(data,g,indent=3)
            g.flush()
            g.seek(0)
            
            output = BytesIO()
            zip = ZipFile(output,'w')
            zip.writestr("Data_Twitter.csv",f.getvalue())
            zip.writestr("Data_Twitter.JSON",g.getvalue())
            zip.close()
            response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=Data_Twitter.zip'
            
            return response
           
            
        elif '0' and '2' in allfile:
            #return csv and xls
            print('csv and xls')
            f = StringIO()    
            
            for row in twitterdata:
                f.write(row.name+','+row.tweet+','+str(row.date)+','+str(row.Retweet_user)+','+row.hashtag)
                f.write('\n')
                
            f.flush()
            f.seek(0)
            
            g = StringIO()
            wb = Workbook()
            ws = wb.active
            
            for row in twitterdata:
                ws.append([row.name,row.tweet,row.date,row.Retweet_user,row.hashtag])
                
            output = BytesIO()
            zip = ZipFile(output,'w')
            zip.writestr("Data_Twitter.csv",f.getvalue())
            zip.writestr("Data_Twitter.xlsx",save_virtual_workbook(wb))
            zip.close()
            response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=Data_Twitter.zip'
            
            return response
        else :
            #return json xls
            print('json and xls')
            
            #make json file
            f = StringIO()
            data =[]
            for row in twitterdata:
                data1 = {}
                data1["NAME"]=row.name
                data1["TWEET"] = row.tweet
                data1["RETWEET_USER"] = row.Retweet_user
                data1["HASHTAGS"] = row.hashtag
                data1["DATE"]= str(row.date)
                data.append(data1)
            json.dump(data,f,indent=3)
            f.flush()
            f.seek(0)
            
            #make xls file
            g = StringIO()
            wb = Workbook()
            ws = wb.active
            
            for row in twitterdata:
                ws.append([row.name,row.tweet,row.date,row.Retweet_user,row.hashtag])
                
            output = BytesIO()
            zip = ZipFile(output,'w')
            zip.writestr("Data_Twitter.csv",f.getvalue())
            zip.writestr("Data_Twitter.xlsx",save_virtual_workbook(wb))
            zip.close()
            response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=Data_Twitter.zip'
            
            return response
        
    elif panjang == 0:
        
    ##asli
        print('do nothing')
        return HttpResponseRedirect('../search2')
         
    else:
        if '0' in allfile:
            f = StringIO()    
            writer = csv.writer(f)
        
            for row in twitterdata:
                writer.writerow([row.name,row.tweet,row.date,row.Retweet_user,row.hashtag])
            f.flush()
            f.seek(0)
            response = HttpResponse(FileWrapper(f), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=Data_twitter.csv'
            return response
        elif '1' in allfile:
            f = StringIO()
            data =[]
            for row in twitterdata:
                data1 = {}
                data1["NAME"]=row.name
                data1["TWEET"] = row.tweet
                data1["RETWEET_USER"] = row.Retweet_user
                data1["HASHTAGS"] = row.hashtag
                data1["DATE"]= str(row.date)
                data.append(data1)
            json.dump(data,f,indent=3)
            f.flush()
            f.seek(0)
            response = HttpResponse(FileWrapper(f), content_type='js/json')
            response['Content-Disposition'] = 'attachment; filename=Data_twitter.json'
            return response
        else:
            f = StringIO()
            wb = Workbook()
            ws = wb.active
            
            for row in twitterdata:
                ws.append([row.name,row.tweet,row.date,row.Retweet_user,row.hashtag])
                
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Data_twitter.xlsx'
            
            return response
        
def downloadfile1(allfile ,facebookdata):
    panjang = len(allfile)
    
    
    
    if panjang == 3:
        ##do something
        f = StringIO()    
        wb = Workbook()
        ws = wb.active
            
        for row in facebookdata:
            f.write(row.name+','+row.status+','+str(row.like)+','+str(row.comment)+','+str(row.share))
            f.write('\n')
            ws.append([row.name,row.status,row.like,row.comment,row.share])
            
        f.flush()
        f.seek(0)
        
        ##make json
        g = StringIO()
        data =[]
        for row in facebookdata:
            data1 = {}
            data1["NAME"]=row.name
            data1["STATUS"] = row.status
            data1["LIKE"] = row.like
            data1["COMMENT"] = row.comment
            data1["SHARE"]= row.share
            data.append(data1)
        json.dump(data,g,indent=3)
        g.flush()
        g.seek(0)
        
        
        
        output = BytesIO()
        zip = ZipFile(output,'w')
        zip.writestr("Data_facebook.csv",f.getvalue())
        zip.writestr("Data_facebook.JSON",g.getvalue())
        zip.writestr("Data_facebook.xlsx",save_virtual_workbook(wb))
        zip.close()
        response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
        response['Content-Disposition'] = 'attachment; filename=Data_facebook.zip'
        
        return response
    
    elif panjang == 2:
        
        ##read list
        if '0' and '1' in allfile:
            ##return csv and json
            print('csv and jsn')
            #make csv
            f = StringIO()    
            
            for row in facebookdata:
                f.write(row.name+','+row.status+','+str(row.like)+','+str(row.comment)+','+str(row.share))
                f.write('\n')
                
            f.flush()
            f.seek(0)
            
            ##make json
            g = StringIO()
            data =[]
            for row in facebookdata:
                data1 = {}
                data1["NAME"]=row.name
                data1["STATUS"] = row.status
                data1["LIKE"] = row.like
                data1["COMMENT"] = row.comment
                data1["SHARE"]= row.share
                data.append(data1)
            json.dump(data,g,indent=3)
            g.flush()
            g.seek(0)
            
            output = BytesIO()
            zip = ZipFile(output,'w')
            zip.writestr("Data_facebook.csv",f.getvalue())
            zip.writestr("Data_facebook.JSON",g.getvalue())
            zip.close()
            response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=Data_facebook.zip'
            
            return response
           
            
        elif '0' and '2' in allfile:
            #return csv and xls
            print('csv and xls')
            f = StringIO()    
            
            for row in facebookdata:
                f.write(row.name+','+row.status+','+str(row.like)+','+str(row.comment)+','+str(row.share))
                f.write('\n')
                
            f.flush()
            f.seek(0)
            
            g = StringIO()
            wb = Workbook()
            ws = wb.active
            
            for row in facebookdata:
                ws.append([row.name,row.status,row.like,row.comment,row.share])
                
            output = BytesIO()
            zip = ZipFile(output,'w')
            zip.writestr("Data_facebook.csv",f.getvalue())
            zip.writestr("Data_facebook.xlsx",save_virtual_workbook(wb))
            zip.close()
            response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=Data_facebook.zip'
            
            return response
        else :
            #return json xls
                        
            #make json file
            f = StringIO()
            data =[]
            for row in facebookdata:
                data1 = {}
                data1["NAME"]=row.name
                data1["TWEET"] = row.tweet
                data1["RETWEET_USER"] = row.Retweet_user
                data1["HASHTAGS"] = row.hashtag
                data1["DATE"]= str(row.date)
                data.append(data1)
            json.dump(data,f,indent=3)
            f.flush()
            f.seek(0)
            
            #make xls file
            g = StringIO()
            wb = Workbook()
            ws = wb.active
            
            for row in facebookdata:
                ws.append([row.name,row.status,row.like,row.comment,row.share])
                
            output = BytesIO()
            zip = ZipFile(output,'w')
            zip.writestr("Data_facebook.csv",f.getvalue())
            zip.writestr("Data_facebook.xlsx",save_virtual_workbook(wb))
            zip.close()
            response = HttpResponse(output.getvalue(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=Data_facebook.zip'
            
            return response
        
    elif panjang == 0:
        
    ##asli
        print('do nothing')
        return HttpResponseRedirect('../search2')
         
    else:
        if '0' in allfile:
            f = StringIO()    
            writer = csv.writer(f)
        
            for row in facebookdata:
                writer.writerow([row.name,row.status,row.like,row.comment,row.share])
            f.flush()
            f.seek(0)
            response = HttpResponse(FileWrapper(f), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=Data_facebook.csv'
            return response
        elif '1' in allfile:
            f = StringIO()
            data =[]
            for row in facebookdata:
                data1 = {}
                data1["NAME"]=row.name
                data1["STATUS"] = row.status
                data1["LIKE"] = row.like
                data1["COMMENT"] = row.comment
                data1["SHARE"]= row.share
                data.append(data1)
            json.dump(data,f,indent=3)
            f.flush()
            f.seek(0)
            response = HttpResponse(FileWrapper(f), content_type='js/json')
            response['Content-Disposition'] = 'attachment; filename=Data_facebook.json'
            return response
        else:
            f = StringIO()
            wb = Workbook()
            ws = wb.active
            
            for row in facebookdata:
                ws.append([row.name,row.status,row.like,row.comment,row.share])
                
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Data_facebook.xlsx'
            
            return response
        
def gettweet1():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_temp = os.path.join(BASE_DIR, 'db.sqlite3')
    
    
    conn = sqlite3.connect(db_temp)
    cursor = conn.cursor()
    a = cursor.execute("SELECT TWEET from social_media_crawling_TwitterCrawl LIMIT 15")
    c = []
    for row in a:
        c.append(row[0])
    conn.close()
    
    return c

def getPost():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_temp = os.path.join(BASE_DIR, 'db.sqlite3')
    
    
    conn = sqlite3.connect(db_temp)
    cursor = conn.cursor()
    a = cursor.execute("SELECT status from social_media_crawling_FacebookCrawl")
    c = []
    for row in a:
        c.append(row[0])
    conn.close()
    
    return c


####################################### JST ###################################
def toJst(statusMI, vocabSize, stopwords,arrCorpus):
    arrStopwords = StopwordsIDDB.objects.values_list('kataStopword', flat=True)
    remove = string.punctuation
    remove = remove.replace("#", "")
    remove = remove.replace("@", "")
    
    dictKata = {} #berisi TF dari masing2 kata
    indexDoc = 0
    dictData = {} #berisi raw dokumen
    dfjKata = {} #berisi banyak dokumen yang memuat suatu kata
    numDocs = len(arrCorpus)


    #Buat data untuk MI dan Formalisasi database
    arrDataMI = []
    formalisasi = FormalisasiKataDB.objects.values_list('kataInformal', 'kataFormal')
    kataFormalisasi = {}
    for i in range(0, len(formalisasi)):
        kataFormalisasi[str(formalisasi[i][0])] = str(formalisasi[i][1])

    #Menyimpan data mentahan dan formalisasi untuk ektraksi MI
    for reader in arrCorpus:
        #reader = ''.join(reader)
        line = str(reader).lower()
        baris = line.split()

        if (len(baris) > 0):
            dictData[indexDoc] = line
            indexDoc += 1

        if (len(baris) > 0):
            kalimat = ""
            for x in range(0, len(baris)):
                if baris[x] in kataFormalisasi.keys():
                    baris[x] = kataFormalisasi[baris[x]]
                kalimat = kalimat + " " + baris[x]
            arrDataMI.append(kalimat)

    #Hitung TF dari masing2 kata
    for line in arrDataMI:
        line = line.translate(str.maketrans('', '', remove)).lower()
        baris = line.split()
        if (len(baris) > 0):
            #TF untuk unigram
            for kata in baris:
                if kata not in dictKata.keys():
                    dictKata[kata] = 1
                    dfjKata[kata] = 0
                else:
                    dictKata[kata] += 1
            #TF untuk bigram
            for i in range(0, (len(baris) - 1)):
                kata = baris[i] + " " + baris[i + 1]
                if kata not in dictKata.keys():
                    dictKata[kata] = 1
                    dfjKata[kata] = 0
                else:
                    dictKata[kata] += 1

    for line in arrDataMI:
        line = line.translate(str.maketrans('', '', remove)).lower()
        baris = line.split()
        if (len(baris) > 0):
            for i in range(0, len(baris) - 1):
                kata = str(''.join(baris[i] + " " + baris[i + 1]))
                baris.append(kata)
        # Menghitung dfj
        for kata in dictKata.keys():
            if kata in baris:
                if (dfjKata[kata] == 0):
                    dfjKata[kata] = 1
                else:
                    dfjKata[kata] += 1

    # Inisialisasi dan hitung tf-idf
    tfidfKata = {}

    # Cek stopwords
#         stopwords = request.POST['stopwords']
    if (stopwords == 'yes'):
        for kata in arrStopwords:
            if kata in dictKata.keys():
                # logging.warning(str(kata))
                del dictKata[kata]

    for kata in dictKata.keys():
        # logging.warning(kata)
        if (dfjKata[kata] == numDocs):
            n = 0
        else:
            n = 1
        tfidfKata[kata] = dictKata[kata] * np.log(numDocs / (dfjKata[kata] + n))
        # logging.warning(str(kata) +" : "+str(tfidfKata[kata]))

    # arrKata = sorted(dictKata, key=dictKata.__getitem__, reverse=True)
    arrKata = sorted(tfidfKata, key=tfidfKata.__getitem__, reverse=True)

    # file.close()
    #arrKata = sorted(dictKata, key=dictKata.__getitem__, reverse=True)

    w = 0
    kata = {}
    for word in arrKata:
        kata[w] = word
        w += 1
#         vocabSize = int(request.POST['vocabSize'])
        if (w > vocabSize):
            w = vocabSize
    
    #logging.warning(str(kata[0]))
    kalimat = str(kata[0])
    for i in range(1, w):
        kalimat = kalimat +","+kata[i]
    if(statusMI == 'yes'):
        statusMI = True
    else:
        statusMI = False
    
    return dictData, kata, indexDoc,w, kalimat, statusMI


################################  PREPROSS ###############################
def downloadprepros(masukan):
    f = StringIO()    
    for a in masukan:
        print('manja',a)
        f.write(a)
        f.write('\n')
    f.flush()
    f.seek(0)
    
    response = HttpResponse(FileWrapper(f), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=test.csv'
    return response

def todlnn():
    Topik = listfolder()
    FeatX = ['1  Bag of Word','2  TF Binary','3  TF-IDF','4  Bigram'] #pilihan fitur extrac
    kelasSentimen = ['2 Kelas Sentimen - (positif atau negatif)','3 Kelas Sentimen - (positif, negatif, atau netral)'] #pilihan jumlah kelas sentimen

    return Topik, FeatX, kelasSentimen